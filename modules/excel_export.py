"""
Excel Export Module — RoadMate Pro
Exports all user trips to a formatted .xlsx file using openpyxl.
"""
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill():
    return PatternFill('solid', fgColor='1A73E8')


def _alt_fill():
    return PatternFill('solid', fgColor='EEF2FF')


def generate_excel(trips_with_costs, user):
    """Generate a professional Excel workbook and return as BytesIO buffer."""
    if not HAS_OPENPYXL:
        return None, "openpyxl not installed. Run: pip install openpyxl"

    wb = openpyxl.Workbook()

    # ── SHEET 1: Trip Summary ──────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Trip Summary'

    # Title row
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f'RoadMate Pro — Trip Report | {user["full_name"]} | {datetime.utcnow().strftime("%d %b %Y")}'
    title_cell.font      = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill      = PatternFill('solid', fgColor='0D47A1')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = [
        'Trip #', 'Source', 'Destination', 'Distance (km)',
        'Travel Days', 'Budget Band', 'Fuel Cost (Rs)',
        'Toll Cost (Rs)', 'Accommodation (Rs)',
        'Total Cost (Rs)', 'Cost/km (Rs)', 'Date Planned'
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color='FFFFFF', size=10)
        c.fill      = _header_fill()
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[2].height = 22

    # Data rows
    for i, t in enumerate(trips_with_costs, 1):
        costs = t.get('costs', {})
        row   = i + 2
        values = [
            i,
            t['source'],
            t['destination'],
            int(t['distance_km']),
            t['travel_days'],
            t['accommodation_budget'].title(),
            round(costs.get('fuel_cost', 0), 2),
            round(costs.get('toll_cost', 0), 2),
            round(costs.get('accommodation_cost', 0), 2),
            round(costs.get('total_cost', 0), 2),
            round(costs.get('cost_per_km', 0), 2),
            t['created_at'][:10],
        ]
        fill = _alt_fill() if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = fill
            c.border    = _border()
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col in (7, 8, 9, 10, 11):
                c.number_format = '#,##0.00'

    # Totals row
    if trips_with_costs:
        tr = len(trips_with_costs) + 3
        ws.cell(tr, 1, 'TOTALS').font = Font(bold=True)
        ws.cell(tr, 4, sum(int(t['distance_km']) for t in trips_with_costs)).font = Font(bold=True)
        for col, key in [(7,'fuel_cost'),(8,'toll_cost'),(9,'accommodation_cost'),(10,'total_cost')]:
            c = ws.cell(tr, col,
                        round(sum(t['costs'].get(key, 0) for t in trips_with_costs if t.get('costs')), 2))
            c.font         = Font(bold=True)
            c.fill         = PatternFill('solid', fgColor='E8F5E9')
            c.number_format = '#,##0.00'
            c.border        = _border()

    # Column widths
    col_widths = [7, 16, 16, 14, 12, 13, 14, 14, 17, 15, 13, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A3'

    # ── SHEET 2: Carbon Footprint ──────────────────────────────────────────
    ws2 = wb.create_sheet('Carbon Footprint')
    ws2.merge_cells('A1:E1')
    h = ws2['A1']
    h.value      = '🌱 Carbon Footprint Estimate — RoadMate Pro'
    h.font       = Font(bold=True, size=13, color='FFFFFF')
    h.fill       = PatternFill('solid', fgColor='2E7D32')
    h.alignment  = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 26

    for col, hdr in enumerate(['Route', 'Fuel Used (L)', 'CO2 Emitted (kg)', 'Trees to Offset', 'Date'], 1):
        c = ws2.cell(2, col, hdr)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor='388E3C')
        c.alignment = Alignment(horizontal='center')
        c.border = _border()

    CO2_PER_LITRE = 2.31
    total_co2 = 0
    for i, t in enumerate(trips_with_costs, 1):
        costs = t.get('costs', {})
        litres = costs.get('fuel_required_liters', 0)
        co2    = round(litres * CO2_PER_LITRE, 2)
        trees  = round(co2 / 21.7, 1)
        total_co2 += co2
        row = i + 2
        fill = PatternFill('solid', fgColor='F1F8E9') if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for col, val in enumerate([
            f"{t['source']} → {t['destination']}",
            litres, co2, trees, t['created_at'][:10]
        ], 1):
            c = ws2.cell(row, col, val)
            c.fill = fill
            c.border = _border()
            c.alignment = Alignment(horizontal='center')

    # Total CO2
    tr2 = len(trips_with_costs) + 3
    ws2.cell(tr2, 1, 'TOTAL').font = Font(bold=True)
    c = ws2.cell(tr2, 3, round(total_co2, 2))
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor='C8E6C9')
    c = ws2.cell(tr2, 4, round(total_co2 / 21.7, 1))
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor='C8E6C9')

    for i, w in enumerate([28, 14, 16, 15, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── SHEET 3: Cost Breakdown ────────────────────────────────────────────
    ws3 = wb.create_sheet('Cost Breakdown')
    ws3.merge_cells('A1:D1')
    h3 = ws3['A1']
    h3.value     = 'Cost Category Breakdown'
    h3.font      = Font(bold=True, size=13, color='FFFFFF')
    h3.fill      = PatternFill('solid', fgColor='E65100')
    h3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 26

    for col, hdr in enumerate(['Category', 'Total Amount (Rs)', 'Percentage', 'Trips Count'], 1):
        c = ws3.cell(2, col, hdr)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='F57C00')
        c.alignment = Alignment(horizontal='center')
        c.border = _border()

    fuel  = sum(t['costs'].get('fuel_cost', 0) for t in trips_with_costs if t.get('costs'))
    toll  = sum(t['costs'].get('toll_cost', 0) for t in trips_with_costs if t.get('costs'))
    hotel = sum(t['costs'].get('accommodation_cost', 0) for t in trips_with_costs if t.get('costs'))
    grand = fuel + toll + hotel or 1
    n     = len(trips_with_costs)

    for row, (cat, val) in enumerate([('Fuel', fuel), ('Tolls', toll), ('Accommodation', hotel)], 3):
        pct = round(val / grand * 100, 1)
        for col, v in enumerate([cat, round(val, 2), f'{pct}%', n], 1):
            c = ws3.cell(row, col, v)
            c.border = _border()
            c.alignment = Alignment(horizontal='center')
    ws3.cell(6, 1, 'GRAND TOTAL').font = Font(bold=True)
    c = ws3.cell(6, 2, round(grand, 2))
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor='FFF3E0')

    for i, w in enumerate([18, 18, 12, 12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Done
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, None
